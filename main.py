import os.path
import openpyxl
import urllib.request as urllib2

from bs4 import BeautifulSoup
from datetime import datetime

def ask_input():
    print('Enter a url from a vehicle on the CarFax website only. e.g. https://www.carfax.com/vehicle/4S3GTAK65H3723107'
          '\nEnter to stop.\n')

    urls = []
    while True:
        url = input()
        if not url:
            break
        if "carfax.com" not in url:
            print(url, ' is not a valid url')
            continue
        urls.append(url)

    return scrape_carfax(urls)

def scrape_carfax(urls, verbose=True):
    data = []
    col_headers = []
    result = []

    for url in urls:
        cars = []
        # querying of the target website and return the html to the variable 'page'
        page = urllib2.urlopen(url)

        # parses the html using beautiful soup and stores it
        soup = BeautifulSoup(page, 'html.parser')

        # dive into the tags to find the name
        car_name = soup.find('div', attrs={'class': 'vehicle-title-container'}).find('h1').text.strip()
        cars.append(car_name)

        price = soup.find('div', attrs={'class': 'vehicle-info-details-price'}).text.strip()
        cars.append(price)

        # get the details numbers
        details = soup.select('div[class=vehicle-info-details]')

        # get the header titles
        detail_titles = soup.find_all('div', class_='vehicle-info-details-title')

        # add the details to the list
        for item in details:
            cars.append(item.text)

        stock_num = soup.find('div', attrs={'class': 'test-auto-stock'}).text.strip()
        cars.append(stock_num)

        cars.append(str(datetime.now()))
        cars.append(str(url))

        # adding the headers for each column
        if not col_headers:
            col_headers.append("Car Name")
            for item2 in detail_titles:
                col_headers.append(item2.text)
            col_headers.append("Update Date")
            col_headers.append("URL")

        data.append(cars)

    result.append(col_headers)
    result.append(data)
    if verbose:
        print(result)
    return result

def append_worksheet(data):
    if not os.path.exists('carEvalsAutomated.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Car Evaluations"
    else:
        wb = openpyxl.load_workbook('carEvalsAutomated.xlsx')
        ws = wb.get_sheet_by_name('Car Evaluations')

    actual_data = data[1]
    headers = data[0]
    i = 0

    for column in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=column)
        if not cell.value:
            cell.value = headers[i]
        i += 1

    row = 1
    i = 1
    while True:
        if ws.cell(row=i, column=1).value is None:
            print(i)
            row = i
            break
        i += 1

    for row, array in enumerate(actual_data, start=row - 1):
        for col, value in enumerate(array):
            cell = ws.cell(row=row + 1, column=col + 1)
            cell.value = value

    wb.save('carEvalsAutomated.xlsx')


data = ask_input()
append_worksheet(data)
